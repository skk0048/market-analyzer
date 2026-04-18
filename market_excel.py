"""
в•”в•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•—
в•‘  EXCEL FORMATTER  v5.1  вЂ” market_excel.py                                 в•‘
в•‘  Shared styling for India & US reports                                     в•‘
в•‘                                                                            в•‘
в•‘  Key fixes v5.1:                                                           в•‘
в•‘   вЂў Data starts at row 1 (title), row 2 (info), row 3 (HEADER)            в•‘
в•‘   вЂў Freeze pane on row 3 only (freeze_panes = "A4")                       в•‘
в•‘   вЂў AutoFilter on header row 3                                             в•‘
в•‘   вЂў Column A width minimised to 6                                          в•‘
в•‘   вЂў TV_Symbol styled in every stock sheet                                  в•‘
в•‘   вЂў Sector-group header rows in Top Picks sheets (coloured dividers)       в•‘
в•љв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ђв•ќ
"""
import numpy as np, pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule

# в”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђ
#  COLOR PALETTE
# в”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђ
TAB_COLORS = {
    "рџ“ё Market Snapshot":    "0A1628",
    "рџЏ­ Sector Strength":    "004D40",
    "рџ”„ Sector Rotation":    "1A237E",
    "рџЏ­ Industry Rotation":  "006064",
    "рџ“Љ Market Breadth":     "1B5E20",
    "рџ“€ Sector Performance": "4A148C",
    "рџ“Љ Stock Strength":     "0D47A1",
    "рџЏ† Top Picks - Buy":    "1B5E20",
    "рџ”ґ Top Picks - Sell":   "B71C1C",
    "рџ“ђ Chart Patterns":     "880E4F",
    "рџЋЇ Trade Setups":       "E65100",
    "рџ“‹ RS Sleeve Lists":    "1A3A5C",
}
HDR_COLORS = {
    "рџ“ё Market Snapshot":    "0D2137",
    "рџЏ­ Sector Strength":    "00695C",
    "рџ”„ Sector Rotation":    "1A237E",
    "рџЏ­ Industry Rotation":  "006064",
    "рџ“Љ Market Breadth":     "2E7D32",
    "рџ“€ Sector Performance": "4A148C",
    "рџ“Љ Stock Strength":     "0D47A1",
    "рџЏ† Top Picks - Buy":    "1B5E20",
    "рџ”ґ Top Picks - Sell":   "B71C1C",
    "рџ“ђ Chart Patterns":     "880E4F",
    "рџЋЇ Trade Setups":       "BF360C",
}
SIG_COLORS = {
    "Strong Buy": ("FFFFFF","006B3C"),
    "Buy":        ("000000","C8E6C9"),
    "Sell":       ("FFFFFF","B71C1C"),
    "Neutral":    ("5D4037","FFF9C4"),
    "WAIT":       ("5D4037","FFF9C4"),
    "BUY":        ("FFFFFF","1B5E20"),
    "SELL":       ("FFFFFF","C62828"),
}
TREND_MAP = {
    "Strong Bullish":("C8E6C9","1B5E20"), "Bullish":("DCEDC8","33691E"),
    "Neutral":("FFF9C4","5D4037"),
    "Bearish":("FFCDD2","B71C1C"), "Strong Bearish":("EF9A9A","7F0000"),
    "Mixed":("FFE0B2","E65100"),
    "в†‘ Bullish":("C8E6C9","1B5E20"), "в†“ Bearish":("FFCDD2","B71C1C"),
    "в†’ Recovering":("DCEDC8","33691E"), "в†’ Pulling Back":("FFE0B2","E65100"),
    "BULLISH":("C8E6C9","1B5E20"), "MIXED":("FFF9C4","5D4037"),
    "BEARISH":("FFCDD2","B71C1C"),
}
ZONE_MAP = {"Bullish":("C8E6C9","1B5E20"), "Neutral":("FFF9C4","5D4037"), "Bearish":("FFCDD2","B71C1C")}
THIN = Side(style="thin", color="D0D0D0")

def _F(h): return PatternFill("solid", fgColor=h)
def _fn(bold=False, color="111111", size=10, italic=False):
    return Font(name="Arial", bold=bold, color=color, size=size, italic=italic)
def _al(h="center", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
def _bd(): return Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def _w(ws, row, col, val, bold=False, fg="111111", bg=None, size=10, h="center", wrap=False):
    c = ws.cell(row, col, val)
    c.font = _fn(bold=bold, color=fg, size=size)
    c.alignment = _al(h, wrap=wrap)
    c.border = _bd()
    if bg: c.fill = _F(bg)
    return c

def _color_cell(cell, col_name, val):
    """Apply coloring based on column name and value."""
    col = col_name.lower().strip()
    # Signal columns
    if col in ("signal","enhanced","sec_signal","action","rs_signal"):
        v = str(val or "")
        if v in SIG_COLORS:
            fg, bg = SIG_COLORS[v]
            cell.fill = _F(bg); cell.font = _fn(bold=True, color=fg)
        return
    # Trend
    if "trend" in col:
        for k,(bg,fg) in TREND_MAP.items():
            if k.lower() in str(val or "").lower():
                cell.fill = _F(bg); cell.font = _fn(bold=True, color=fg); return
        return
    # Zone
    if "zone" in col:
        v = str(val or "")
        if v in ZONE_MAP:
            bg,fg = ZONE_MAP[v]; cell.fill = _F(bg); cell.font = _fn(bold=True, color=fg)
        return
    # Tick вњ“/вњ—
    if col.startswith("abv_") or "beats" in col:
        v = str(val or "")
        if v == "вњ“": cell.fill = _F("C8E6C9"); cell.font = _fn(bold=True, color="1B5E20")
        elif v == "вњ—": cell.fill = _F("FFCDD2"); cell.font = _fn(color="B71C1C")
        return
    # Chart pattern
    if col == "chart_pattern":
        v = str(val or "")
        if v.startswith("рџџў"): cell.fill = _F("E3F2FD"); cell.font = _fn(bold=True, color="1565C0")
        elif v.startswith("рџ”ґ"): cell.fill = _F("FFEBEE"); cell.font = _fn(bold=True, color="C62828")
        return
    # Direction
    if col == "direction":
        v = str(val or "")
        if v=="BULLISH": cell.fill=_F("E3F2FD"); cell.font=_fn(bold=True,color="1565C0")
        elif v=="BEARISH": cell.fill=_F("FFEBEE"); cell.font=_fn(bold=True,color="C62828")
        return
    # RSI
    if col in ("rsi_14","rsi"):
        if isinstance(val,(int,float)) and not np.isnan(val):
            if val>=60: cell.fill=_F("C8E6C9"); cell.font=_fn(bold=True,color="1B5E20")
            elif val>=50: cell.fill=_F("F1F8E9"); cell.font=_fn(color="33691E")
            elif val<40: cell.fill=_F("FFCDD2"); cell.font=_fn(color="B71C1C")
        return
    # SMA Score
    if col == "sma_score":
        pal={4:("C8E6C9","1B5E20"),3:("DCEDC8","33691E"),2:("FFF9C4","5D4037"),
             1:("FFE0B2","E65100"),0:("FFCDD2","B71C1C")}
        try:
            v=int(float(val))
            if v in pal: bg,fg=pal[v]; cell.fill=_F(bg); cell.font=_fn(bold=(v==4),color=fg)
        except: pass
        return
    # Breadth % bar (0-100)
    if col in ("rs22%","rs55%","rsi50%","abvsma20%","abvsma50%","abvsma100%","abvsma200%",
               "1m_score","3m_score","6m_score"):
        if isinstance(val,(int,float)) and not np.isnan(val):
            if val>=60: cell.fill=_F("C8E6C9"); cell.font=_fn(bold=True,color="1B5E20")
            elif val>=40: cell.fill=_F("FFF9C4"); cell.font=_fn(color="5D4037")
            else: cell.fill=_F("FFCDD2"); cell.font=_fn(color="B71C1C")
        return
    # D/E
    if col in ("d/e","de"):
        if isinstance(val,(int,float)) and not np.isnan(val):
            if val<0.5: cell.fill=_F("C8E6C9"); cell.font=_fn(color="1B5E20")
            elif val<1: cell.fill=_F("FFF9C4"); cell.font=_fn(color="5D4037")
            elif val>=2: cell.fill=_F("FFCDD2"); cell.font=_fn(color="B71C1C")
        return
    # TV_Symbol вЂ” styled distinctly
    if col == "tv_symbol":
        cell.fill=_F("E8F5E9"); cell.font=_fn(bold=True,color="1B5E20",size=9)
        return
    # Action
    if col == "action":
        v=str(val or "")
        if v=="BUY": cell.fill=_F("1B5E20"); cell.font=_fn(bold=True,color="FFFFFF")
        elif v=="SELL": cell.fill=_F("B71C1C"); cell.font=_fn(bold=True,color="FFFFFF")
        elif v=="WAIT": cell.fill=_F("FFF9C4"); cell.font=_fn(bold=True,color="5D4037")
        return
    # General % (positive=green, negative=red)
    pct_set={"chg_1d%","chg_5d%","avg_chg_1d%","avg_chg_5d%","rs_22d%","rs_55d%","rs_120d%","rs_252d%",
             "1m%","3m%","6m%","12m%","ytd%","rs_1m%","rs_3m%","rs_6m%","rs_12m%",
             "from_52w_high%","rs_score","total_score","sales_qoq%","sales_yoy%",
             "pat_qoq%","pat_yoy%","margin%","roe%","sec_rs55%",
             "rs_22d_idx%","rs_55d_idx%","rs_22d_sec%","rs_55d_sec%",
             "rs_120d_idx%","rs_252d_idx%",f"ret_22d%",f"ret_55d%"}
    if col in pct_set or col.endswith("%"):
        if isinstance(val,(int,float)) and not np.isnan(val):
            bold=abs(val)>5
            if val>0: cell.fill=_F("C8E6C9"); cell.font=_fn(bold=bold,color="1B5E20")
            elif val<0: cell.fill=_F("FFCDD2"); cell.font=_fn(bold=bold,color="B71C1C")
        return

# в”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђ
#  CORE SHEET WRITER
#  Layout:  Row 1 = Title (merged, dark bg)
#           Row 2 = Info subtitle (italic, grey)
#           Row 3 = HEADER  в†ђ AutoFilter + Freeze below here
#           Row 4+ = Data
# в”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђ
def write_sheet(ws, df, sheet_name, title="", freeze_row=4):
    if df is None or df.empty:
        ws.cell(1,1,"No data available for this section.")
        return

    hdr_bg = HDR_COLORS.get(sheet_name,"0D2137")
    n_cols = len(df.columns)

    # в”Ђв”Ђ Row 1: Title в”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђ
    if title:
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
        c = ws.cell(1,1,title)
        c.fill = _F(hdr_bg); c.font = _fn(bold=True,color="FFFFFF",size=13)
        c.alignment = _al("left"); ws.row_dimensions[1].height = 28

    # в”Ђв”Ђ Row 2: Info в”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђ
    ts = datetime.now().strftime("%d-%b-%Y %H:%M")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=min(n_cols,10))
    info_cell = ws.cell(2,1,f"Generated: {ts}  |  {len(df)} rows")
    info_cell.font = _fn(italic=True, color="777777", size=9)
    info_cell.alignment = _al("left")
    ws.row_dimensions[2].height = 14

    # в”Ђв”Ђ Row 3: Header в”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђ
    headers = list(df.columns)
    for j, h in enumerate(headers, 1):
        c = ws.cell(3, j, h)
        c.fill = _F(hdr_bg); c.font = _fn(bold=True, color="FFFFFF", size=10)
        c.alignment = _al("center", wrap=True); c.border = _bd()
    ws.row_dimensions[3].height = 24

    # AutoFilter on header row
    ws.auto_filter.ref = f"A3:{get_column_letter(n_cols)}3"

    # Freeze below header (A4)
    ws.freeze_panes = f"A{freeze_row}"

    hdr_idx = {h: j+1 for j,h in enumerate(headers)}

    # Columns that are text в†’ left-align
    _LEFT_COLS = {'symbol','tv_symbol','company','company name','name',
                  'sector','industry','chart_pattern','notes','setup_desc',
                  'signal_type','strategy','trend'}

    # в”Ђв”Ђ Rows 4+: Data в”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђ
    for i, (_, row_data) in enumerate(df.iterrows()):
        r = i + 4
        alt_bg = "F7F9FC" if i % 2 == 0 else "FFFFFF"
        for j, col in enumerate(headers, 1):
            val = row_data[col]
            if isinstance(val, float) and np.isnan(val): val = ""
            c = ws.cell(r, j, val)
            halign = "left" if col.lower() in _LEFT_COLS else "center"
            c.alignment = _al(halign)
            c.border = _bd()
            c.fill = _F(alt_bg)
            c.font = _fn(size=10)
            _color_cell(c, col, val)
        ws.row_dimensions[r].height = 17

    # в”Ђв”Ђ Column widths в”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђ
    # Col A (Rank/index) в†’ narrow
    ws.column_dimensions["A"].width = 6
    for col_obj in ws.iter_cols(min_row=3, max_row=3):
        j = col_obj[0].column
        col_letter = get_column_letter(j)
        header_val = str(col_obj[0].value or "")
        max_len = max(
            len(header_val),
            max((len(str(ws.cell(r, j).value or "")) for r in range(4, ws.max_row+1)), default=8)
        )
        # Special widths
        if header_val.lower() in ("rank","sec_rank","valid","adv","dec","h_day","sma_score","fin_score"):
            ws.column_dimensions[col_letter].width = 7
        elif header_val.lower() == "tv_symbol":
            ws.column_dimensions[col_letter].width = 22
        elif header_val.lower() in ("symbol",):
            ws.column_dimensions[col_letter].width = 12
        elif header_val.lower() in ("company","company name","name"):
            ws.column_dimensions[col_letter].width = 28
        elif header_val.lower() in ("sector","industry"):
            ws.column_dimensions[col_letter].width = 18
        elif header_val.lower() == "chart_pattern":
            ws.column_dimensions[col_letter].width = 24
        elif header_val.lower() in ("setup_type","notes","1m_zone","3m_zone","6m_zone"):
            ws.column_dimensions[col_letter].width = min(max_len+3, 28)
        elif header_val.lower() == "trend":
            ws.column_dimensions[col_letter].width = 16
        else:
            ws.column_dimensions[col_letter].width = min(max_len+2, 16)

    ws.sheet_view.showGridLines = False

# в”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђ
#  TOP PICKS SHEET WRITER  (sector-group header dividers)
# в”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђ
def write_top_picks_sheet(ws, df, sheet_name, title="", is_sell=False):
    if df is None or df.empty:
        ws.cell(1,1,"No data available."); return
    if "Message" in df.columns:
        ws.cell(1,1,df["Message"].iloc[0]); return

    hdr_bg = HDR_COLORS.get(sheet_name,"0D2137")
    # Remove internal helper cols
    display_cols = [c for c in df.columns if not c.startswith("_")]
    n_cols = len(display_cols)

    # Row 1: Title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    c = ws.cell(1,1,title)
    c.fill = _F(hdr_bg); c.font = _fn(bold=True,color="FFFFFF",size=13)
    c.alignment = _al("left"); ws.row_dimensions[1].height = 28

    # Row 2: Info
    ts = datetime.now().strftime("%d-%b-%Y %H:%M")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=min(n_cols,10))
    ws.cell(2,1,f"Generated: {ts}  |  Sectors ranked {'strongestв†’weakest' if not is_sell else 'weakestв†’strongest'} by RS_55d%")
    ws.cell(2,1).font = _fn(italic=True,color="777777",size=9)
    ws.cell(2,1).alignment = _al("left"); ws.row_dimensions[2].height = 14

    # Row 3: Header
    for j,h in enumerate(display_cols,1):
        c = ws.cell(3,j,h)
        c.fill=_F(hdr_bg); c.font=_fn(bold=True,color="FFFFFF",size=10)
        c.alignment=_al("center",wrap=True); c.border=_bd()
    ws.row_dimensions[3].height=24
    ws.auto_filter.ref = f"A3:{get_column_letter(n_cols)}3"
    ws.freeze_panes = "A4"

    # Data with sector-group divider rows
    r = 4
    prev_sec = None
    for i, (_, row_data) in enumerate(df.iterrows()):
        sec = row_data.get("Sector","")
        # Insert sector divider when sector changes
        if sec != prev_sec:
            sec_rs = row_data.get("Sec_RS55%", np.nan)
            sec_sig= row_data.get("Sec_Signal","")
            sec_rank=row_data.get("Sec_Rank","")
            div_text = f"  в–ё  #{sec_rank}  {sec}   {sec_sig}   RS_55d: {sec_rs:+.1f}%" if not np.isnan(sec_rs) else f"  в–ё  #{sec_rank}  {sec}   {sec_sig}"
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=n_cols)
            dc = ws.cell(r,1,div_text)
            div_bg = "C8E6C9" if sec_sig=="Buy" else ("FFCDD2" if sec_sig=="Sell" else "FFF9C4")
            div_fg = "1B5E20" if sec_sig=="Buy" else ("B71C1C" if sec_sig=="Sell" else "5D4037")
            dc.fill=_F(div_bg); dc.font=_fn(bold=True,color=div_fg,size=10)
            dc.alignment=_al("left"); dc.border=_bd()
            ws.row_dimensions[r].height=20; r+=1; prev_sec=sec

        alt_bg="F0FFF0" if not is_sell else "FFF5F5"
        alt_bg=alt_bg if i%2==0 else "FFFFFF"
        _LEFT_COLS2 = {'symbol','tv_symbol','company','company name','name',
                       'sector','industry','chart_pattern','notes','setup_desc','signal_type','strategy','trend'}
        for j,col in enumerate(display_cols,1):
            val=row_data.get(col,"")
            if isinstance(val,float) and np.isnan(val): val=""
            c=ws.cell(r,j,val)
            halign = "left" if col.lower() in _LEFT_COLS2 else "center"
            c.alignment=_al(halign); c.border=_bd(); c.fill=_F(alt_bg); c.font=_fn(size=10)
            _color_cell(c,col,val)
        ws.row_dimensions[r].height=17; r+=1

    # Column widths (same logic)
    ws.column_dimensions["A"].width=6
    for col_obj in ws.iter_cols(min_row=3,max_row=3):
        j=col_obj[0].column; ltr=get_column_letter(j); h=str(col_obj[0].value or "").lower()
        if h in ("rank","sec_rank"): ws.column_dimensions[ltr].width=7
        elif h=="tv_symbol": ws.column_dimensions[ltr].width=22
        elif h=="symbol": ws.column_dimensions[ltr].width=12
        elif h in ("company","company name"): ws.column_dimensions[ltr].width=28
        elif h in ("sector","industry"): ws.column_dimensions[ltr].width=18
        elif h=="chart_pattern": ws.column_dimensions[ltr].width=24
        elif h=="trend": ws.column_dimensions[ltr].width=16
        else: ws.column_dimensions[ltr].width=min(len(str(col_obj[0].value or ""))+3,16)
    ws.sheet_view.showGridLines=False

# в”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђ
#  RS SLEEVE LIST SHEET WRITER  (multi-section: A / B / C + legend)
# в”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђ

_SLEEVE_DIVIDER_BG   = "1A3A5C"   # dark navy for sleeve header rows
_SLEEVE_LEGEND_BG    = "263238"   # charcoal for methodology rows
_SLEEVE_A_BG         = "E3F2FD"   # light blue rows  вЂ” Large Cap
_SLEEVE_B_BG         = "E8F5E9"   # light green rows вЂ” Mid-Large
_SLEEVE_C_BG         = "FFF8E1"   # light amber rows вЂ” Small-Mid
_SLEEVE_USA_BG       = "F3E5F5"   # light purple     вЂ” US sleeves


def write_rs_sleeve_sheet(ws, df, market="INDIA"):
    """
    Write the RS Sleeve / Smallcase Action List sheet.
    Divider rows (Rank starts with 'в”Ѓв”Ѓв”Ѓ') get special dark header styling.
    Data rows get alternating tier-colour backgrounds.
    """
    ws.sheet_view.showGridLines = False

    title = (
        f"рџ“‹  RS SLEEVE LISTS  [{market}]  вЂ”  Smallcase / MF-style Momentum Portfolios"
        f"  |  A = Large Cap (Monthly)  В·  B = Growth (Fortnightly)  В·  C = Aggressive (Weekly)"
    )
    # Row 1: merged title
    n_cols = max(len(df.columns), 15)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    tc = ws.cell(1, 1, title)
    tc.fill = _F(_SLEEVE_DIVIDER_BG)
    tc.font = _fn(bold=True, color="00E5FF", size=12)
    tc.alignment = _al("left")
    ws.row_dimensions[1].height = 26

    # Row 2: column headers
    headers = list(df.columns)
    hdr_display = {
        "Rank": "Rank", "Symbol": "Symbol", "Company": "Company / Description",
        "Sector": "Sector", "Industry": "Industry", "Price": "Price / Detail",
        "Sleeve_RS": "Sleeve RS Score", "RS_22d_Idx%": "RS 22d%",
        "RS_55d_Idx%": "RS 55d%", "RS_120d_Idx%": "RS 120d%", "RS_252d_Idx%": "RS 252d%",
        "Signal": "Signal", "Enhanced": "Enhanced", "RSI_14": "RSI",
        "Supertrend": "Supertrend", "SL_Buy%": "SL Buy%", "SL_Grade": "SL Grade",
        "SL_Buy_Price": "SL Price", "MST_Signal": "MST", "LST_Signal": "LST",
        "RS30_Signal": "RS30", "Sales_YoY%": "Sales YoY%", "PAT_YoY%": "PAT YoY%",
        "ROE%": "ROE%", "Mkt_Cap_B": "Mkt Cap B", "Chart_Pattern": "Pattern",
    }
    for j, h in enumerate(headers, 1):
        c = ws.cell(2, j, hdr_display.get(h, h))
        c.fill = _F(_SLEEVE_DIVIDER_BG)
        c.font = _fn(bold=True, color="FFFFFF", size=10)
        c.alignment = _al("center", wrap=True)
        c.border = _bd()
    ws.row_dimensions[2].height = 22
    ws.freeze_panes = "A3"

    # Detect which sleeve a row belongs to (for background colour)
    cur_sleeve = "A"; cur_bg = _SLEEVE_A_BG
    _sleeve_bg_map = {
        "A": _SLEEVE_A_BG, "B": _SLEEVE_B_BG, "C": _SLEEVE_C_BG,
        "US_A": _SLEEVE_USA_BG, "US_B": _SLEEVE_B_BG, "US_C": _SLEEVE_C_BG,
    }
    # Regime banner background colours
    _REGIME_BG = {"BULL": "1B5E20", "CAUTION": "E65100", "BEAR": "B71C1C"}

    for i, (_, row_data) in enumerate(df.iterrows()):
        r = i + 3
        rank_val = str(row_data.get("Rank", "") or "")

        is_divider = rank_val.startswith("в”Ѓв”Ѓв”Ѓ")
        is_legend  = is_divider and "METHOD" in rank_val.upper()
        is_regime  = is_divider and "MARKET REGIME" in rank_val.upper()
        is_blank   = all(str(v or "").strip() == "" for v in row_data.values)

        if is_blank:
            ws.row_dimensions[r].height = 7
            continue

        if is_divider:
            # Update current sleeve colour
            for key in _sleeve_bg_map:
                if f"SLEEVE {key}" in rank_val or f"SLEEVE {key.replace('_','')}" in rank_val:
                    cur_sleeve = key
                    cur_bg = _sleeve_bg_map.get(key, _SLEEVE_A_BG)
                    break

            # Choose background based on type
            if is_regime:
                sym_val = str(row_data.get("Symbol", "") or "")
                if "BEAR" in sym_val and "CAUTION" not in sym_val:
                    div_bg = _REGIME_BG["BEAR"]
                elif "CAUTION" in sym_val:
                    div_bg = _REGIME_BG["CAUTION"]
                else:
                    div_bg = _REGIME_BG["BULL"]
            elif is_legend:
                div_bg = _SLEEVE_LEGEND_BG
            else:
                div_bg = _SLEEVE_DIVIDER_BG

            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=n_cols)
            parts = [str(row_data.get(h, "") or "")
                     for h in ["Rank", "Symbol", "Company", "Sector", "Industry", "Price", "Sleeve_RS"]
                     if str(row_data.get(h, "") or "").strip()]
            div_text = "   |   ".join(parts[:6])
            dc = ws.cell(r, 1, div_text)
            dc.fill = _F(div_bg)
            dc.font = _fn(bold=True, color="00E5FF" if is_legend else "FFFFFF", size=10)
            dc.alignment = _al("left")
            dc.border = _bd()
            ws.row_dimensions[r].height = 22 if is_regime else 20
            continue

        # Normal data row
        alt_bg = cur_bg if i % 2 == 0 else "FFFFFF"
        _LEFT = {'symbol','company','sector','industry','chart_pattern','signal',
                 'enhanced','mst_signal','lst_signal','rs30_signal','supertrend'}
        for j, col in enumerate(headers, 1):
            val = row_data.get(col, "")
            if isinstance(val, float) and np.isnan(val): val = ""
            c = ws.cell(r, j, val)
            halign = "left" if col.lower() in _LEFT else "center"
            c.alignment = _al(halign)
            c.border = _bd()
            c.fill = _F(alt_bg)
            c.font = _fn(size=10)
            _color_cell(c, col, val)
        ws.row_dimensions[r].height = 17

    # Column widths
    col_widths = {
        "Rank": 6, "Symbol": 12, "Company": 28, "Sector": 16,
        "Industry": 18, "Price": 10, "Sleeve_RS": 12,
        "RS_22d_Idx%": 9, "RS_55d_Idx%": 9, "RS_120d_Idx%": 10, "RS_252d_Idx%": 10,
        "Avg_Turnover": 12, "Daily_Std%": 10, "Equal_Wt%": 9, "ATR_Wt%": 9,
        "Signal": 10, "Enhanced": 12, "RSI_14": 7, "Supertrend": 10,
        "SL_Buy%": 8, "SL_Grade": 8, "SL_Buy_Price": 10,
        "MST_Signal": 7, "LST_Signal": 7, "RS30_Signal": 7,
        "Sales_YoY%": 9, "PAT_YoY%": 9, "ROE%": 7, "Mkt_Cap_B": 9,
        "Chart_Pattern": 22,
    }
    for j, col in enumerate(headers, 1):
        ltr = get_column_letter(j)
        ws.column_dimensions[ltr].width = col_widths.get(col, 12)


# в”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђ
#  MAIN WORKBOOK BUILDER  вЂ” 11 sheets in fixed order
# в”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђ
def build_workbook(market, snapshot_df, sector_str_df, sector_rot_df,
                   industry_rot_df, breadth_df, sector_perf_df, stock_str_df,
                   top_buy_df, top_sell_df, chart_pat_df, trade_df, output_path,
                   dashboard_df=None, sleeve_df=None):
    wb = Workbook(); wb.remove(wb.active)
    run_time = datetime.now().strftime("%d %b %Y  %H:%M")

    # 0. Dashboard (Sheet 1)
    ws=wb.create_sheet("рџ“‹ Dashboard")
    if dashboard_df is not None and not dashboard_df.empty:
        write_dashboard_sheet(ws, dashboard_df, market)
    else:
        ws.cell(1,1,"Dashboard вЂ” run market_india.py or market_usa.py to populate.")
    ws.sheet_properties.tabColor = "0A1628"
    ws.sheet_view.showGridLines = False

    # 1. Market Snapshot
    ws=wb.create_sheet("рџ“ё Market Snapshot")
    write_sheet(ws, snapshot_df, "рџ“ё Market Snapshot",
                f"рџ“ё  MARKET SNAPSHOT  [{market}]  вЂ”  {run_time}")

    # 2. Sector Strength
    ws=wb.create_sheet("рџЏ­ Sector Strength")
    write_sheet(ws, sector_str_df, "рџЏ­ Sector Strength",
                "рџЏ­  SECTOR STRENGTH  вЂ”  Ranked by RS vs Benchmark")

    # 3. Sector Rotation
    ws=wb.create_sheet("рџ”„ Sector Rotation")
    write_sheet(ws, sector_rot_df, "рџ”„ Sector Rotation",
                "рџ”„  SECTOR ROTATION  вЂ”  Breadth & Zone Metrics")

    # 4. Industry Rotation
    ws=wb.create_sheet("рџЏ­ Industry Rotation")
    write_sheet(ws, industry_rot_df, "рџЏ­ Industry Rotation",
                "рџЏ­  INDUSTRY ROTATION  вЂ”  Breadth & Zone Metrics")

    # 5. Market Breadth
    ws=wb.create_sheet("рџ“Љ Market Breadth")
    write_sheet(ws, breadth_df, "рџ“Љ Market Breadth",
                f"рџ“Љ  MARKET BREADTH  [{market}]  вЂ”  Adv/Dec + % Above MAs")

    # 6. Sector Performance
    ws=wb.create_sheet("рџ“€ Sector Performance")
    write_sheet(ws, sector_perf_df, "рџ“€ Sector Performance",
                "рџ“€  SECTOR PERFORMANCE  вЂ”  1M / 3M / 6M / YTD Returns")

    # 7. Stock Strength
    ws=wb.create_sheet("рџ“Љ Stock Strength")
    write_sheet(ws, stock_str_df, "рџ“Љ Stock Strength",
                "рџ“Љ  STOCK STRENGTH  вЂ”  RS + Technicals + Financials (all-in-one)")
    # RS Score colour scale on Stock Strength
    if not stock_str_df.empty and "RS_Score" in stock_str_df.columns:
        rs_col_idx = list(stock_str_df.columns).index("RS_Score")+1
        rs_ltr = get_column_letter(rs_col_idx)
        ws.conditional_formatting.add(
            f"{rs_ltr}4:{rs_ltr}{len(stock_str_df)+3}",
            ColorScaleRule(start_type="min",start_color="FF4444",
                           mid_type="num",mid_value=0,mid_color="FFEB84",
                           end_type="max",end_color="00C853"))

    # 8. Top Picks - Buy
    ws=wb.create_sheet("рџЏ† Top Picks - Buy")
    write_top_picks_sheet(ws, top_buy_df, "рџЏ† Top Picks - Buy",
                          "рџЏ†  TOP BUY STOCKS  вЂ”  Strongest Sector в†’ Weakest  |  в‰Ґ5 Stocks per Sector",
                          is_sell=False)

    # 9. Top Picks - Sell
    ws=wb.create_sheet("рџ”ґ Top Picks - Sell")
    write_top_picks_sheet(ws, top_sell_df, "рџ”ґ Top Picks - Sell",
                          "рџ”ґ  TOP SELL STOCKS  вЂ”  Weakest Sector в†’ Strongest  |  в‰Ґ5 Stocks per Sector",
                          is_sell=True)

    # 10. Chart Patterns
    ws=wb.create_sheet("рџ“ђ Chart Patterns")
    write_sheet(ws, chart_pat_df, "рџ“ђ Chart Patterns",
                "рџ“ђ  CHART PATTERNS  вЂ”  Last 14 Days Only  |  Entry / Stop / Target / RR")

    # 11. Trade Setups
    ws=wb.create_sheet("рџЋЇ Trade Setups")
    write_sheet(ws, trade_df, "рџЋЇ Trade Setups",
                "рџЋЇ  TRADE SETUPS  вЂ”  BUY / SELL / WAIT  |  RS + Fundamentals Combined Score")

    # 12. RS Sleeve / Smallcase Action List
    if sleeve_df is not None and not sleeve_df.empty:
        ws=wb.create_sheet("рџ“‹ RS Sleeve Lists")
        write_rs_sleeve_sheet(ws, sleeve_df, market)

    # Tab colours
    for ws_obj in wb.worksheets:
        for key, color in TAB_COLORS.items():
            if key in ws_obj.title:
                ws_obj.sheet_properties.tabColor = color; break
        ws_obj.sheet_view.showGridLines = False

    wb.save(output_path)
    print(f"  рџ’ѕ Saved: {output_path}")
    return output_path


# в”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђ
#  DASHBOARD SHEET WRITER  (added v5.1)
# в”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђ

def write_dashboard_sheet(ws, df, market="INDIA"):
    """Write the рџ“‹ Dashboard as a two-column info sheet."""
    ws.sheet_view.showGridLines = False

    # Col widths
    ws.column_dimensions["A"].width = 46
    ws.column_dimensions["B"].width = 100

    SECTION_BG  = "1F3864"   # dark blue for section headers
    SECTION_FG  = "FFFFFF"
    DATA_BG     = "F4F6FB"   # light blue-grey for data rows
    ALT_BG      = "FFFFFF"
    GREEN_BG    = "C8E6C9"; GREEN_FG = "1B5E20"
    RED_BG      = "FFCDD2";  RED_FG  = "B71C1C"
    AMBER_BG    = "FFF9C4";  AMBER_FG= "5D4037"
    TV_BG       = "E8F5E9";  TV_FG   = "1B5E20"

    # Title row
    ws.merge_cells("A1:B1")
    t = ws.cell(1, 1, f"рџ“‹  FundaTechno Market Analysis  [{market}]")
    t.fill = _F("0A1628"); t.font = _fn(bold=True, color="00E5FF", size=14)
    t.alignment = _al("left"); ws.row_dimensions[1].height = 32

    row = 2
    for i, (_, dr) in enumerate(df.iterrows()):
        k = str(dr["Key"]   or "")
        v = str(dr["Value"] or "")
        is_section = k.startswith("в”Ђв”Ђ")
        is_blank   = (k == "")
        is_tv      = k.startswith("TV") or k == "All Buy" or k == "Strong Buy" or k in ("MST Buy","LST Buy","RS30 Buy")

        if is_blank:
            ws.row_dimensions[row].height = 8
            row += 1
            continue

        if is_section:
            ws.merge_cells(f"A{row}:B{row}")
            c = ws.cell(row, 1, k)
            c.fill = _F(SECTION_BG); c.font = _fn(bold=True, color=SECTION_FG, size=11)
            c.alignment = _al("left"); ws.row_dimensions[row].height = 22
        else:
            alt = (i % 2 == 0)
            bg  = DATA_BG if alt else ALT_BG
            c1  = ws.cell(row, 1, k)
            c2  = ws.cell(row, 2, v)
            c1.fill = _F(bg); c1.font = _fn(bold=True, color="333333", size=10)
            c1.alignment = _al("left")
            c2.fill = _F(bg); c2.font = _fn(size=9, color="222222")
            c2.alignment = _al("left", wrap=True)

            # Special value coloring
            if is_tv:
                c1.fill = _F(TV_BG); c1.font = _fn(bold=True, color=TV_FG, size=10)
                c2.fill = _F(TV_BG); c2.font = _fn(bold=True, color=TV_FG, size=9)
            elif k in ("в­ђ Strong Buy (all 5 peer filters)", "вњ… Buy", "MST Buy", "LST Buy", "RS30 Buy",
                       "Grade A  в‰¤3%  (Ideal)", "Grade B  3-5% (Good)"):
                try:
                    n = int(v)
                    if n > 0:
                        c2.fill = _F(GREEN_BG); c2.font = _fn(bold=True, color=GREEN_FG, size=10)
                except: pass
            elif k in ("рџ”ґ Sell",):
                try:
                    n = int(v)
                    if n > 0:
                        c2.fill = _F(RED_BG); c2.font = _fn(bold=True, color=RED_FG, size=10)
                except: pass
            ws.row_dimensions[row].height = 18
        row += 1

    ws.freeze_panes = "A2"


# в”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђ
#  EXTENDED COLOR HANDLER  вЂ” adds new v5.1 columns
# в”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђв”Ђ

_ORIG_color_cell = _color_cell

def _color_cell(cell, col_name, val):
    """Extended version: handles all v5.1 new columns."""
    col = col_name.lower().strip()

    # MST / LST / RS30 signals
    if col in ("mst_signal","lst_signal","rs30_signal"):
        v = str(val or "")
        if v == "Buy":
            cell.fill = _F("C8E6C9"); cell.font = _fn(bold=True, color="1B5E20")
        elif v == "Watch":
            cell.fill = _F("FFF9C4"); cell.font = _fn(bold=True, color="5D4037")
        elif v == "Neutral":
            cell.fill = _F("F5F5F5"); cell.font = _fn(color="888888")
        return

    # Supertrend
    if col == "supertrend":
        v = str(val or "")
        if v == "Buy":
            cell.fill = _F("C8E6C9"); cell.font = _fn(bold=True, color="1B5E20")
        elif v == "Sell":
            cell.fill = _F("FFCDD2"); cell.font = _fn(bold=True, color="B71C1C")
        elif v == "N/A":
            cell.fill = _F("F5F5F5"); cell.font = _fn(color="AAAAAA", italic=True)
        return

    # SL Grade
    if col == "sl_grade":
        pal={"A":("C8E6C9","1B5E20"),"B":("DCEDC8","33691E"),
             "C":("FFF9C4","5D4037"),"D":("FFE0B2","E65100"),"F":("FFCDD2","B71C1C")}
        v = str(val or "")
        if v in pal:
            bg,fg = pal[v]; cell.fill=_F(bg); cell.font=_fn(bold=(v=="A"),color=fg)
        return

    # SL%  (lower = green, higher = red вЂ” INVERTED from normal %)
    if col in ("sl_buy%","sl%","sl_sell%","sl_buy_pct","sl_sell_pct"):
        if isinstance(val,(int,float)) and not np.isnan(val):
            if   val <= 3:  cell.fill=_F("C8E6C9"); cell.font=_fn(bold=True,color="1B5E20")
            elif val <= 5:  cell.fill=_F("DCEDC8"); cell.font=_fn(color="33691E")
            elif val <= 8:  cell.fill=_F("FFF9C4"); cell.font=_fn(color="5D4037")
            elif val <= 12: cell.fill=_F("FFE0B2"); cell.font=_fn(color="E65100")
            else:           cell.fill=_F("FFCDD2"); cell.font=_fn(color="B71C1C")
        return

    # SL Price / Swing levels вЂ” just format as number, no color
    if col in ("sl_price","sl_buy_price","sl_sell_price","swing_low_20d","swing_high_20d","tp1_price","tp2_price"):
        if isinstance(val,(int,float)) and not np.isnan(val):
            cell.font = _fn(bold=False, color="333333", size=10)
        return

    # TP % (always positive вЂ” standard green if >0)
    if col in ("tp1%","tp2%"):
        if isinstance(val,(int,float)) and not np.isnan(val) and val > 0:
            cell.fill=_F("C8E6C9"); cell.font=_fn(bold=True,color="1B5E20")
        return

    # RR Ratio
    if col in ("rr_t1","rr_t2"):
        if isinstance(val,(int,float)) and not np.isnan(val):
            if   val >= 3: cell.fill=_F("C8E6C9"); cell.font=_fn(bold=True,color="1B5E20")
            elif val >= 2: cell.fill=_F("FFF9C4"); cell.font=_fn(color="5D4037")
            else:          cell.fill=_F("FFCDD2"); cell.font=_fn(color="B71C1C")
        return

    # Breakout flag
    if col == "breakout_20d":
        v = str(val or "")
        if v == "вњ“": cell.fill=_F("C8E6C9"); cell.font=_fn(bold=True,color="1B5E20")
        elif v == "вњ—": cell.fill=_F("F5F5F5"); cell.font=_fn(color="AAAAAA")
        return

    # Signal_Type (colourful labels in Trade Setups)
    if col in ("signal_type","setup_desc"):
        v = str(val or "")
        if any(x in v for x in ["RS30","LST","MST"]):
            cell.fill=_F("E3F2FD"); cell.font=_fn(bold=True,color="1565C0")
        elif "Strong" in v:
            cell.fill=_F("C8E6C9"); cell.font=_fn(bold=True,color="006B3C")
        elif "Sell" in v or "рџ”ґ" in v:
            cell.fill=_F("FFCDD2"); cell.font=_fn(bold=True,color="B71C1C")
        elif "Watch" in v or "вЏі" in v:
            cell.fill=_F("FFF9C4"); cell.font=_fn(color="5D4037")
        return

    # ATR weight % вЂ” higher = more weight = neutral green
    if col in ("atr_wt%", "equal_wt%"):
        if isinstance(val, (int, float)) and not np.isnan(val):
            if val >= 10: cell.fill=_F("C8E6C9"); cell.font=_fn(bold=True,color="1B5E20")
            elif val >= 5: cell.fill=_F("F1F8E9"); cell.font=_fn(color="33691E")
            else:          cell.fill=_F("FFF9C4"); cell.font=_fn(color="5D4037")
        return

    # Daily std % вЂ” lower = greener (tighter = better)
    if col == "daily_std%":
        if isinstance(val, (int, float)) and not np.isnan(val):
            if   val <= 1.5: cell.fill=_F("C8E6C9"); cell.font=_fn(bold=True,color="1B5E20")
            elif val <= 2.5: cell.fill=_F("FFF9C4"); cell.font=_fn(color="5D4037")
            else:            cell.fill=_F("FFCDD2"); cell.font=_fn(color="B71C1C")
        return

    # Avg turnover вЂ” higher = greener
    if col == "avg_turnover":
        if isinstance(val, (int, float)) and not np.isnan(val):
            if   val >= 100: cell.fill=_F("C8E6C9"); cell.font=_fn(bold=True,color="1B5E20")
            elif val >= 20:  cell.fill=_F("F1F8E9"); cell.font=_fn(color="33691E")
            elif val >= 5:   cell.fill=_F("FFF9C4"); cell.font=_fn(color="5D4037")
            else:            cell.fill=_F("FFCDD2"); cell.font=_fn(color="B71C1C")
        return

    # Fall through to original handler
    _ORIG_color_cell(cell, col_name, val)
